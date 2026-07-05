import os
import sys
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
import re
from sqlalchemy.orm import Session

# Setup python path
HERE = Path(__file__).resolve().parent.parent.parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))

from services.quotation_service import generate_quotation_draft, QuotationDraft
from services.ai_provider import AiProviderConfigPayload
from services.tenant_ai_settings import resolve_ai_config_for_request
from storage.db import get_db
from fastapi.responses import StreamingResponse
import io
import openpyxl

router = APIRouter(tags=["Quotation"])

class QuotationRequest(BaseModel):
    query: str = Field(description="고장 증상 및 현상")
    charger_type: str = Field(default="급속", description="충전기 구분: 급속 / 완속")
    tenant_id: str = Field(default="default_tenant", description="테넌트 ID")
    ai_config: AiProviderConfigPayload | None = None

class ExportPartItem(BaseModel):
    part_name: str
    spec: str
    qty: int
    unit_price: int
    category: str

class QuotationExportRequest(BaseModel):
    query: str
    symptom_summary: str
    likely_cause: str
    parts: list[ExportPartItem]
    dispatch_fee: int
    labor_fee: int

@router.post("/quotation/draft", response_model=QuotationDraft)
async def create_quotation_draft(req: QuotationRequest, db: Session = Depends(get_db)):
    from services.billing_metering import (
        FEATURE_AI_GENERATION,
        check_quota,
        record_usage,
    )
    import asyncio

    # FastAPI는 기본적으로 동기 함수를 스레드풀에서 실행하지만,
    # I/O 바운드 작업(LLM 호출)이 명확하므로 async로 정의하여 명시성을 높입니다.
    # 실제 성능 향상을 위해서는 generate_quotation_draft 내부도 async/await 구조로 바꾸는 것이 이상적입니다.

    tenant_id = (req.tenant_id or "default_tenant").strip()
    check_quota(tenant_id, FEATURE_AI_GENERATION)
    try:
        ai_config = resolve_ai_config_for_request(db, tenant_id, req.ai_config)
    except Exception as cfg_exc:
        print(f"AI config load failed, using env defaults: {cfg_exc}")
        ai_config = None

    try:
        # 비동기적으로 실행 (실제 함수가 async가 아니더라도 이벤트 루프 블로킹 방지)
        loop = asyncio.get_running_loop()
        draft = await loop.run_in_executor(
            None, generate_quotation_draft, req.query, req.charger_type, ai_config
        )
        is_faq = draft.symptom_summary.startswith("FAQ:")

        try:
            record_usage(
                tenant_id,
                FEATURE_AI_GENERATION,
                model_name="faq-shortcut" if is_faq else "hybrid",
                shortcut=is_faq,
            )
        except Exception as usage_exc:
            print(f"Usage metering failed (draft still returned): {usage_exc}")
        return draft
    except HTTPException:
        raise
    except FileNotFoundError as e:
        raise HTTPException(status_code=503, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI 견적서 생성 실패: {str(e)}")

@router.post("/quotation/export")
async def export_quotation_excel(req: QuotationExportRequest):
    try:
        def sanitize_filename(filename: str) -> str:
            """파일 이름으로 사용할 수 없는 문자를 제거하거나 대체합니다."""
            return re.sub(r'[\\/*?:"<>|]', "_", filename)


        template_path = HERE / "assets" / "quotation_template.xlsx"
        if not template_path.is_file():
            raise FileNotFoundError(f"견적서 템플릿 파일을 찾을 수 없습니다: {template_path}")
            
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 1. Fill Header Information
        # 원본 query를 사용하여 수신자 정보의 정확도를 높입니다.
        ws.cell(row=2, column=2).value = f"수  신   : '{req.query}' 건 담당 엔지니어 귀하\n주  소   : 현장 점검 대상지"
        
        # 2. Fill Parts and technical fees into the table (starts at row=9, max 8 rows: row 9 to 16)
        all_items = []
        
        # Add parts
        for i, part in enumerate(req.parts, 1):
            all_items.append({
                "no": i,
                "name": part.part_name,
                "spec": part.spec,
                "qty": part.qty,
                "price": part.unit_price,
                "note": part.category
            })
            
        # Add dispatch fee (if greater than 0)
        idx = len(all_items) + 1
        if req.dispatch_fee > 0:
            all_items.append({
                "no": idx,
                "name": "출장 교통비",
                "spec": "기술 서비스료",
                "qty": 1,
                "price": req.dispatch_fee,
                "note": "-"
            })
            idx += 1
            
        # Add labor fee (if greater than 0)
        if req.labor_fee > 0:
            all_items.append({
                "no": idx,
                "name": "작업 공임비",
                "spec": "기술 서비스료",
                "qty": 1,
                "price": req.labor_fee,
                "note": "-"
            })
            
        # Total supply value calculation
        parts_total = sum(p.unit_price * p.qty for p in req.parts)
        supply_value = parts_total + req.dispatch_fee + req.labor_fee
        
        # Fill table rows dynamically
        # The template has one placeholder row at row 9.
        # We will insert new rows for additional items.
        start_row = 9
        if len(all_items) > 1:
            ws.insert_rows(start_row + 1, amount=len(all_items) - 1)

        for idx, item in enumerate(all_items):
            current_row = start_row + idx
            ws.cell(row=current_row, column=2).value = item["no"]
            ws.cell(row=current_row, column=3).value = item["name"]
            ws.cell(row=current_row, column=6).value = item["spec"]
            ws.cell(row=current_row, column=9).value = "EA"
            ws.cell(row=current_row, column=10).value = item["qty"]
            ws.cell(row=current_row, column=12).value = item["price"]
            ws.cell(row=current_row, column=14).value = item["price"] * item["qty"]
            ws.cell(row=current_row, column=18).value = item["note"]
                
        # 3. Fill Totals
        # 공급가액 총합 (상단, row=7, column=6)
        ws.cell(row=7, column=6).value = supply_value
        
        # 합계 (하단, row=17 in original template, now shifted)
        total_row_idx = start_row + len(all_items) + 7 # 7 is the offset from item table to total row
        ws.cell(row=total_row_idx, column=14).value = supply_value
        
        # 4. Save to bytes stream
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        # Return Excel file
        safe_filename = sanitize_filename(f"견적서_{req.query[:15]}.xlsx")
        headers = {
            'Content-Disposition': f'attachment; filename*=UTF-8\'\'{safe_filename}'
        }
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"엑셀 견적서 생성 실패: {str(e)}")
